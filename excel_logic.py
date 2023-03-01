from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# set column width


def set_column_widths():
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20

# set column names


column_names = ['S.No', 'Name', 'Category', 'Subscribers', 'Average Views',
                'Youtube Channel Link', 'Twitter Account', 'Instagram Account', 'Facebook Page']


def set_column_names():
    for index, column_name in enumerate(column_names):
        ws.cell(row=1, column=index+1).value = column_name
        # ws.cell(row=1, column=index+1).alignment = Alignment(horizontal='center')


set_column_widths()
set_column_names()
